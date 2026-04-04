# -*- coding: utf-8 -*-
"""
ChemEng-Bench 200: Multi-Judge Evaluation v3
FIXES v3: GPT-4o-mini→GPT-4o judge, Gemini responses v2 (truncation fix)
"""
import os, sys, json, time, re, math

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except:
        pass

from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

# ============================================================
#  API Keys
# ============================================================
GPT_KEY      = "YOUR_OPENAI_API_KEY"
CLAUDE_KEY   = "YOUR_ANTHROPIC_API_KEY"
DEEPSEEK_KEY = "YOUR_DEEPSEEK_API_KEY"
GROQ_KEY     = "YOUR_GROQ_API_KEY"
GEMINI_KEY   = "YOUR_GEMINI_API_KEY"
# ============================================================

GROUND_TRUTH_FILE = "ChemEng_Bench_200_GroundTruth.xlsx"

RESPONSE_FILES = {
    "GPT-4o":    "gpt4o_responses.xlsx",
    "DeepSeek":  "deepseek_responses.xlsx",
    "Claude":    "claude_responses.xlsx",
    "Gemini":    "gemini_responses_v2.xlsx",
    "Llama":     "llama_responses.xlsx",
}

PROGRESS_FILE = "eval_multijudge_progress_v3.json"

# ============================================================
# CONFIG — ปรับได้ตามต้องการ
# ============================================================
MAX_RETRY       = 3      # จำนวนครั้งที่ลอง retry ต่อ judge
RETRY_DELAY     = 3      # วินาทีรอระหว่าง retry (exponential)
SLEEP_BETWEEN   = 1.0    # วินาทีรอระหว่าง judge แต่ละตัว
SLEEP_BETWEEN_Q = 0.5    # วินาทีรอระหว่าง prompt แต่ละข้อ
MIN_JUDGES      = 3      # ถ้า prompt ได้ judge น้อยกว่านี้ จะ re-evaluate

# ============================================================
# Weights — ปรับน้ำหนักตรงนี้
# ============================================================
W_ROUGE1  = 0.15
W_ROUGEL  = 0.15
W_COSINE  = 0.20
W_ELEMENT = 0.50

# ============================================================
# ROUGE (local)
# ============================================================
def tokenize(text):
    return re.findall(r'\b\w+\b', text.lower())

def rouge_1(reference, hypothesis):
    ref_t = tokenize(reference); hyp_t = tokenize(hypothesis)
    if not ref_t or not hyp_t: return 0.0
    rc = Counter(ref_t); hc = Counter(hyp_t)
    ov = sum(min(rc[t], hc[t]) for t in hc)
    p = ov/len(hyp_t); r = ov/len(ref_t)
    if p+r == 0: return 0.0
    return round(2*p*r/(p+r), 3)

def lcs_length(x, y):
    m, n = len(x), len(y)
    if m == 0 or n == 0: return 0
    prev = [0]*(n+1); curr = [0]*(n+1)
    for i in range(1, m+1):
        for j in range(1, n+1):
            if x[i-1] == y[j-1]: curr[j] = prev[j-1]+1
            else: curr[j] = max(curr[j-1], prev[j])
        prev, curr = curr, [0]*(n+1)
    return prev[n]

def rouge_l(reference, hypothesis):
    ref_t = tokenize(reference)[:500]; hyp_t = tokenize(hypothesis)[:500]
    if not ref_t or not hyp_t: return 0.0
    lcs = lcs_length(ref_t, hyp_t)
    p = lcs/len(hyp_t); r = lcs/len(ref_t)
    if p+r == 0: return 0.0
    return round(2*p*r/(p+r), 3)

# ============================================================
# Cosine Similarity (local TF-IDF)
# ============================================================
def cosine_sim(ref_text, hyp_text):
    rt = tokenize(ref_text); ht = tokenize(hyp_text)
    if not rt or not ht: return 0.0
    docs = [rt, ht]
    def tfidf(tokens):
        tf = Counter(tokens); mx = max(tf.values()) if tf else 1
        v = {}
        for w, c in tf.items():
            df = sum(1 for d in docs if w in set(d))
            v[w] = (c/mx)*(math.log(3/(df+1))+1)
        return v
    v1 = tfidf(rt); v2 = tfidf(ht)
    words = set(list(v1.keys())+list(v2.keys()))
    dot = sum(v1.get(w,0)*v2.get(w,0) for w in words)
    m1 = math.sqrt(sum(v**2 for v in v1.values()))
    m2 = math.sqrt(sum(v**2 for v in v2.values()))
    if m1 == 0 or m2 == 0: return 0.0
    return round(dot/(m1*m2), 3)

# ============================================================
# Scoring Prompt + Parser
# ============================================================
def make_prompt(prompt, rubric, gt, ai_resp):
    return (
        "You are a chemical engineering professor checking a student answer.\n\n"
        "QUESTION:\n" + prompt + "\n\n"
        "SCORING RUBRIC:\n" + rubric + "\n\n"
        "REFERENCE ANSWER:\n" + gt + "\n\n"
        "STUDENT ANSWER:\n" + ai_resp + "\n\n"
        "Check each rubric criterion. Respond ONLY in JSON:\n"
        '{"found": 5, "total": 7, "missing": ["brief name 1", "brief name 2"]}\n'
    )

def parse_json(text):
    text = text.replace("```json","").replace("```","").strip()
    # หา JSON object ใน text
    m = re.search(r'\{[^{}]*\}', text)
    if m:
        text = m.group()
    
    # ลอง parse ปกติก่อน
    try:
        r = json.loads(text)
        return r.get("found",0), r.get("total",7), r.get("missing",[])
    except json.JSONDecodeError:
        pass
    
    # แก้ Gemini: single quotes → double quotes
    fixed = text.replace("'", '"')
    try:
        r = json.loads(fixed)
        return r.get("found",0), r.get("total",7), r.get("missing",[])
    except json.JSONDecodeError:
        pass
    
    # แก้ Gemini: unquoted keys เช่น {found: 5, total: 7}
    fixed2 = re.sub(r'(\w+)\s*:', r'"\1":', text)
    fixed2 = fixed2.replace("'", '"')
    try:
        r = json.loads(fixed2)
        return r.get("found",0), r.get("total",7), r.get("missing",[])
    except json.JSONDecodeError:
        pass
    
    # Fallback: ดึงตัวเลข found/total ด้วย regex
    found_m = re.search(r'["\']?found["\']?\s*[:=]\s*(\d+)', text, re.IGNORECASE)
    total_m = re.search(r'["\']?total["\']?\s*[:=]\s*(\d+)', text, re.IGNORECASE)
    if found_m and total_m:
        return int(found_m.group(1)), int(total_m.group(1)), []
    
    raise ValueError(f"Cannot parse JSON: {text[:100]}")

# ============================================================
# Grade
# ============================================================
def assign_grade(overall):
    if overall >= 0.7: return "Excellent"
    elif overall >= 0.5: return "Good"
    elif overall >= 0.35: return "Fair"
    else: return "Poor"

# ============================================================
# Judge Initialization (แยกเป็น function เพื่อ re-init ได้)
# ============================================================
def init_judge_gpt():
    if not GPT_KEY or len(GPT_KEY) < 15: return None
    try:
        c = OpenAI(api_key=GPT_KEY)
        c.chat.completions.create(model="gpt-4o",max_tokens=5,
            messages=[{"role":"user","content":"OK"}])
        return c
    except Exception as e:
        print("    FAIL: GPT (" + str(e)[:60] + ")")
        return None

def init_judge_claude():
    if not CLAUDE_KEY or len(CLAUDE_KEY) < 15: return None
    try:
        import anthropic
        c = anthropic.Anthropic(api_key=CLAUDE_KEY)
        c.messages.create(model="claude-sonnet-4-20250514",max_tokens=5,
            messages=[{"role":"user","content":"OK"}])
        return c
    except Exception as e:
        print("    FAIL: Claude (" + str(e)[:60] + ")")
        return None

def init_judge_deepseek():
    if not DEEPSEEK_KEY or len(DEEPSEEK_KEY) < 15: return None
    try:
        c = OpenAI(api_key=DEEPSEEK_KEY, base_url="https://api.deepseek.com")
        c.chat.completions.create(model="deepseek-chat",max_tokens=5,
            messages=[{"role":"user","content":"OK"}])
        return c
    except Exception as e:
        print("    FAIL: DeepSeek (" + str(e)[:60] + ")")
        return None

def init_judge_llama():
    if not GROQ_KEY or len(GROQ_KEY) < 15: return None
    try:
        c = OpenAI(api_key=GROQ_KEY, base_url="https://api.groq.com/openai/v1")
        c.chat.completions.create(model="llama-3.3-70b-versatile",max_tokens=5,
            messages=[{"role":"user","content":"OK"}])
        return c
    except Exception as e:
        print("    FAIL: Llama (" + str(e)[:60] + ")")
        return None

def init_judge_gemini():
    if not GEMINI_KEY or len(GEMINI_KEY) < 15: return None
    try:
        from google import genai
        c = genai.Client(api_key=GEMINI_KEY)
        c.models.generate_content(
            model="gemini-2.5-flash", contents="Say OK",
            config=genai.types.GenerateContentConfig(
                max_output_tokens=100,
                response_mime_type="application/json"))
        return c
    except Exception as e:
        print("    FAIL: Gemini (" + str(e)[:60] + ")")
        return None

JUDGE_INITS = {
    "GPT": init_judge_gpt,
    "Claude": init_judge_claude,
    "DeepSeek": init_judge_deepseek,
    "Llama": init_judge_llama,
    "Gemini": init_judge_gemini,
}

def init_all_judges():
    jlist = {}
    for name, func in JUDGE_INITS.items():
        c = func()
        if c:
            jlist[name] = c
            print("  OK:   " + name)
        else:
            print("  skip: " + name)
    return jlist

# ============================================================
# Call Judge with RETRY + RE-INIT
# ============================================================
def call_judge(jname, jclient, prompt, rubric, gt_answer, ai_resp):
    sp = make_prompt(prompt, rubric, gt_answer, ai_resp)
    
    if jname == "Claude":
        resp = jclient.messages.create(
            model="claude-sonnet-4-20250514", max_tokens=300,
            system="Respond in valid JSON only. No markdown.",
            messages=[{"role":"user","content":sp}])
        return parse_json(resp.content[0].text)
    
    elif jname == "Gemini":
        from google import genai
        resp = jclient.models.generate_content(
            model="gemini-2.5-flash",
            contents="Respond in valid JSON only. No markdown.\n\n" + sp,
            config=genai.types.GenerateContentConfig(
                max_output_tokens=1024,
                response_mime_type="application/json"))
        return parse_json(resp.text)
    
    else:
        model_name = {
            "GPT": "gpt-4o",
            "DeepSeek": "deepseek-chat",
            "Llama": "llama-3.3-70b-versatile",
        }.get(jname, "gpt-4o")
        
        resp = jclient.chat.completions.create(
            model=model_name, max_tokens=300, temperature=0,
            messages=[
                {"role":"system","content":"Respond in valid JSON only. No markdown."},
                {"role":"user","content":sp}
            ])
        return parse_json(resp.choices[0].message.content)

def call_judge_with_retry(jname, jclient, prompt, rubric, gt_answer, ai_resp, judges_dict):
    """เรียก judge พร้อม retry และ re-init ถ้าจำเป็น"""
    for attempt in range(MAX_RETRY):
        try:
            f, t, m = call_judge(jname, jclient, prompt, rubric, gt_answer, ai_resp)
            return f, t, m, jclient  # ส่ง client กลับด้วย (อาจเปลี่ยนหลัง re-init)
        except Exception as ex:
            err_msg = str(ex)[:60]
            if attempt < MAX_RETRY - 1:
                wait = RETRY_DELAY * (2 ** attempt)
                print(f"      RETRY {jname} (attempt {attempt+2}/{MAX_RETRY}) "
                      f"wait {wait}s... [{err_msg}]")
                time.sleep(wait)
                
                # ลอง re-init judge ในรอบ retry สุดท้าย
                if attempt == MAX_RETRY - 2:
                    print(f"      RE-INIT {jname}...")
                    new_client = JUDGE_INITS[jname]()
                    if new_client:
                        jclient = new_client
                        judges_dict[jname] = new_client
                        print(f"      RE-INIT {jname} OK")
                    else:
                        print(f"      RE-INIT {jname} FAILED")
            else:
                return None, None, None, jclient  # ล้มเหลวทุก retry

# ============================================================
# MAIN
# ============================================================
print("=" * 60)
print("  ChemEng-Bench 200: Multi-Judge Evaluation v3")
print("  Features: retry, re-init, per-judge Excel, re-evaluate")
print("=" * 60)

print("\nInitializing judges...")
judges = init_all_judges()

if not judges:
    print("\nERROR: No judges available!"); sys.exit(1)
print(f"\n  Active: {len(judges)} judges => {', '.join(judges.keys())}")

# -----------------------------------------------------------
# Load Ground Truth
# -----------------------------------------------------------
if not os.path.exists(GROUND_TRUTH_FILE):
    print("\nERROR: Not found: " + GROUND_TRUTH_FILE); sys.exit(1)

wb = openpyxl.load_workbook(GROUND_TRUTH_FILE)
ws = wb.active
gt = {}
for r in range(2, ws.max_row+1):
    pid = ws.cell(r,2).value
    if pid:
        gt[pid] = {
            "no": ws.cell(r,1).value, "domain": str(ws.cell(r,3).value or ""),
            "prompt": str(ws.cell(r,4).value or ""), "bloom": ws.cell(r,5).value,
            "diff": ws.cell(r,6).value, "answer": str(ws.cell(r,7).value or ""),
            "rubric": str(ws.cell(r,8).value or ""),
        }
print(f"  Ground truth: {len(gt)}")

# -----------------------------------------------------------
# Load Response Files
# -----------------------------------------------------------
print("\nResponse files:")
available = {}
for name, fp in RESPONSE_FILES.items():
    if os.path.exists(fp):
        wb2 = openpyxl.load_workbook(fp); ws2 = wb2.active
        resp = {}
        for r in range(2, ws2.max_row+1):
            pid = ws2.cell(r,2).value; ai = ws2.cell(r,7).value
            if pid and ai and not str(ai).startswith("ERROR"):
                resp[pid] = str(ai)
        if resp:
            available[name] = resp
            print(f"  FOUND: {name} ({len(resp)})")
    else:
        print(f"  skip:  {name}")

if not available:
    print("\nERROR: No response files!"); sys.exit(1)

# -----------------------------------------------------------
# Load Progress + Identify prompts needing re-evaluation
# -----------------------------------------------------------
done = {}
if os.path.exists(PROGRESS_FILE):
    with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
        done = json.load(f)
    total_done = sum(len(v) for v in done.values())
    
    # นับ prompt ที่ judge น้อยเกินไป
    redo_count = 0
    for mn in done:
        for pid, d in done[mn].items():
            if d.get("num_judges", 0) < MIN_JUDGES:
                redo_count += 1
    print(f"\n  Resuming: {total_done} done, {redo_count} need re-evaluation (judges < {MIN_JUDGES})")

# ============================================================
# EVALUATE
# ============================================================
for model_name, responses in available.items():
    if model_name not in done:
        done[model_name] = {}
    
    # แยก: prompt ใหม่ + prompt ที่ต้อง re-evaluate
    new_pids = [pid for pid in responses if pid in gt and pid not in done[model_name]]
    redo_pids = [pid for pid in responses if pid in gt and pid in done[model_name]
                 and done[model_name][pid].get("num_judges", 0) < MIN_JUDGES]
    
    remaining = new_pids + redo_pids
    
    if not remaining:
        print(f"\n{model_name}: done ({len(done[model_name])})")
        continue
    
    print(f"\n{'='*60}")
    print(f"Evaluating: {model_name} ({len(new_pids)} new + {len(redo_pids)} redo = {len(remaining)})")
    print(f"{'='*60}")
    
    for j, pid in enumerate(remaining):
        g = gt[pid]; ai_resp = responses[pid]; t0 = time.time()
        is_redo = pid in redo_pids
        
        # Local metrics (ใช้ค่าเดิมถ้า redo)
        if is_redo and done[model_name][pid].get("rouge1"):
            r1 = done[model_name][pid]["rouge1"]
            rl = done[model_name][pid]["rougeL"]
            cos = done[model_name][pid]["cosine"]
        else:
            r1 = rouge_1(g["answer"], ai_resp)
            rl = rouge_l(g["answer"], ai_resp)
            cos = cosine_sim(g["answer"], ai_resp)
        
        # Multi-judge element check WITH RETRY
        all_found = []; all_total = []; all_missing = []
        judge_detail = {}
        
        for jname, jclient in list(judges.items()):
            f, t, m, updated_client = call_judge_with_retry(
                jname, jclient, g["prompt"], g["rubric"], g["answer"], ai_resp, judges)
            
            if updated_client != jclient:
                judges[jname] = updated_client
            
            if f is not None:
                judge_detail[jname] = {"found": f, "total": t}
                all_found.append(f); all_total.append(t)
                if m:
                    all_missing.extend(m)
            else:
                judge_detail[jname] = {"found": "-", "total": "-", "error": "FAILED"}
            
            time.sleep(SLEEP_BETWEEN)
        
        # Average across successful judges
        if all_found:
            avg_f = round(sum(all_found)/len(all_found), 1)
            avg_t = round(sum(all_total)/len(all_total), 1)
            epct = round(avg_f/max(avg_t,1)*100, 1)
        else:
            avg_f = 0; avg_t = 7; epct = 0
        
        # Missing elements
        if all_missing:
            um = list(set(x.strip().lower() for x in all_missing))[:5]
            mstr = ", ".join(um)
        else:
            mstr = "ALL FOUND" if all_found else "NO JUDGES"
        
        elapsed = round(time.time()-t0, 1)
        overall = round(W_ROUGE1*r1 + W_ROUGEL*rl + W_COSINE*cos + W_ELEMENT*(epct/100), 3)
        grade = assign_grade(overall)
        nj = len(all_found)
        
        redo_tag = " [REDO]" if is_redo else ""
        print(f"  [{j+1:3d}/{len(remaining)}] {pid}  "
              f"R1={r1} RL={rl} Cos={cos} "
              f"Elem={avg_f}/{avg_t} ({nj}j) => {grade} ({elapsed}s){redo_tag}")
        
        # แสดงคะแนนแยกแต่ละ judge
        for jn, jd in judge_detail.items():
            if "error" in jd and jd.get("found") == "-":
                print(f"      {jn}: FAILED")
            else:
                print(f"      {jn}: {jd['found']}/{jd['total']}")
        
        done[model_name][pid] = {
            "rouge1":r1, "rougeL":rl, "cosine":cos,
            "avg_found":avg_f, "avg_total":avg_t, "elem_pct":epct,
            "overall":overall, "grade":grade, "missing":mstr,
            "time":elapsed, "judges":judge_detail, "num_judges":nj,
        }
        
        with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
            json.dump(done, f, ensure_ascii=False)
        time.sleep(SLEEP_BETWEEN_Q)

# ============================================================
# EXCEL OUTPUT — with per-judge columns
# ============================================================
print(f"\n{'='*60}")
print("Generating Excel with per-judge scores...")

hfont = Font(name='Arial', bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="2E6B2E", end_color="2E6B2E", fill_type="solid")
hfill2 = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # judge header
bdr = Border(left=Side(style='thin'),right=Side(style='thin'),
             top=Side(style='thin'),bottom=Side(style='thin'))
ctr = Alignment(horizontal='center', vertical='center')
wrp = Alignment(horizontal='left', vertical='top', wrap_text=True)
grade_colors = {"Excellent":"D5F5E3","Good":"FEF9E7","Fair":"FDEBD0","Poor":"FADBD8"}

JUDGE_NAMES = ["GPT", "Claude", "DeepSeek", "Llama", "Gemini"]

domain_map = {
    "Process Modeling & Simulation":"MOD","Process Optimization":"OPT",
    "Machine Learning for Chemical Processes":"ML",
    "Process Design & Systems Engineering":"DES",
}

# --- Per-model Excel ---
for mn in available:
    if mn not in done: continue
    wbo = openpyxl.Workbook(); wso = wbo.active; wso.title = mn
    
    # Headers: base columns + per-judge columns
    base_hdrs = ['No.','ID','Domain','Bloom','Diff',
                 'ROUGE-1','ROUGE-L','Cosine',
                 'Elem Avg','Elem %','Overall','Grade']
    judge_hdrs = [f'{jn} (found/total)' for jn in JUDGE_NAMES]
    all_hdrs = base_hdrs + judge_hdrs + ['Missing Elements','Active Judges','Time(s)']
    
    for i, h in enumerate(all_hdrs, 1):
        c = wso.cell(1, i, h); c.font = hfont; c.border = bdr; c.alignment = ctr
        if i <= len(base_hdrs):
            c.fill = hfill
        elif i <= len(base_hdrs) + len(judge_hdrs):
            c.fill = hfill2
        else:
            c.fill = hfill
    
    row = 2
    for pid in sorted(gt.keys()):
        if pid not in done[mn]: continue
        d = done[mn][pid]; g = gt[pid]
        dm = domain_map.get(g["domain"], "?")
        
        col = 1
        # Base columns
        for val in [g["no"], pid, dm, g.get("bloom",""), g.get("diff","")]:
            c = wso.cell(row, col, val); c.border = bdr; c.alignment = ctr; col += 1
        for val in [d["rouge1"], d["rougeL"], d["cosine"]]:
            c = wso.cell(row, col, val); c.border = bdr; c.alignment = ctr; col += 1
        
        elem_str = str(d["avg_found"]) + "/" + str(d["avg_total"])
        c = wso.cell(row, col, elem_str); c.border = bdr; c.alignment = ctr; col += 1
        c = wso.cell(row, col, d["elem_pct"]); c.border = bdr; c.alignment = ctr; col += 1
        c = wso.cell(row, col, d["overall"]); c.border = bdr; c.alignment = ctr; col += 1
        
        gc = wso.cell(row, col, d["grade"]); gc.border = bdr; gc.alignment = ctr
        clr = grade_colors.get(d["grade"], "FFFFFF")
        gc.fill = PatternFill(start_color=clr, end_color=clr, fill_type="solid")
        col += 1
        
        # Per-judge columns
        jdetail = d.get("judges", {})
        for jn in JUDGE_NAMES:
            jd = jdetail.get(jn, {})
            if "error" in jd and jd.get("found") in ["-", None]:
                val = "FAIL"
                c = wso.cell(row, col, val); c.border = bdr; c.alignment = ctr
                c.font = Font(color="FF0000", name='Arial')
            elif jd.get("found") is not None and jd.get("found") != "-":
                val = f"{jd['found']}/{jd['total']}"
                c = wso.cell(row, col, val); c.border = bdr; c.alignment = ctr
            else:
                val = "-"
                c = wso.cell(row, col, val); c.border = bdr; c.alignment = ctr
                c.font = Font(color="999999", name='Arial')
            col += 1
        
        # Missing, Active Judges, Time
        c = wso.cell(row, col, d["missing"]); c.border = bdr; c.alignment = wrp; col += 1
        c = wso.cell(row, col, d.get("num_judges", 0)); c.border = bdr; c.alignment = ctr; col += 1
        c = wso.cell(row, col, d["time"]); c.border = bdr; c.alignment = ctr
        row += 1
    
    # Column widths
    widths = {'A':5,'B':10,'C':6,'D':7,'E':7,'F':9,'G':9,'H':9,
              'I':10,'J':8,'K':9,'L':10}
    for c_letter, w in widths.items():
        wso.column_dimensions[c_letter].width = w
    # Judge columns + remaining
    for i in range(len(base_hdrs)+1, len(all_hdrs)+1):
        letter = openpyxl.utils.get_column_letter(i)
        wso.column_dimensions[letter].width = 14
    
    # Freeze header
    wso.freeze_panes = 'A2'
    
    fn = mn.lower().replace("-","").replace(" ","") + "_evaluation_v3.xlsx"
    wbo.save(fn)
    print(f"  Saved: {fn}")

# --- Summary Excel ---
wbs = openpyxl.Workbook(); wss = wbs.active; wss.title = "Summary"
shdrs = ['Model','ROUGE-1','ROUGE-L','Cosine','Elem %','Overall',
         'Grade','Judges','MOD','OPT','ML','DES']
for i, h in enumerate(shdrs, 1):
    c = wss.cell(1, i, h); c.font = hfont; c.fill = hfill; c.border = bdr; c.alignment = ctr

row = 2
for mn in available:
    if mn not in done: continue
    r1s,rls,coss,elems,ovs=[],[],[],[],[]
    by_dom={"MOD":[],"OPT":[],"ML":[],"DES":[]}; nj=0
    for pid, d in done[mn].items():
        if pid in gt:
            r1s.append(d["rouge1"]); rls.append(d["rougeL"])
            coss.append(d["cosine"]); elems.append(d["elem_pct"])
            ovs.append(d["overall"])
            nj = max(nj, d.get("num_judges", 0))
            dm = domain_map.get(gt[pid]["domain"], "?")
            if dm in by_dom: by_dom[dm].append(d["overall"])
    n = max(len(ovs), 1)
    ar1=round(sum(r1s)/n,3); arl=round(sum(rls)/n,3)
    acs=round(sum(coss)/n,3); ael=round(sum(elems)/n,1)
    aov=round(sum(ovs)/n,3); ag=assign_grade(aov)
    
    wss.cell(row,1,mn).border=bdr
    wss.cell(row,2,ar1).border=bdr; wss.cell(row,2).alignment=ctr
    wss.cell(row,3,arl).border=bdr; wss.cell(row,3).alignment=ctr
    wss.cell(row,4,acs).border=bdr; wss.cell(row,4).alignment=ctr
    wss.cell(row,5,ael).border=bdr; wss.cell(row,5).alignment=ctr
    wss.cell(row,6,aov).border=bdr; wss.cell(row,6).alignment=ctr
    gc=wss.cell(row,7,ag); gc.border=bdr; gc.alignment=ctr
    clr=grade_colors.get(ag,"FFFFFF")
    gc.fill=PatternFill(start_color=clr,end_color=clr,fill_type="solid")
    wss.cell(row,8,nj).border=bdr; wss.cell(row,8).alignment=ctr
    for col,dm in [(9,"MOD"),(10,"OPT"),(11,"ML"),(12,"DES")]:
        v=by_dom[dm]; dd=round(sum(v)/max(len(v),1),3) if v else 0
        wss.cell(row,col,dd).border=bdr; wss.cell(row,col).alignment=ctr
    
    print(f"  {mn.ljust(10)} R1={ar1} RL={arl} Cos={acs} Elem={ael}% => {ag} ({nj}j)")
    row += 1

# --- Judge Agreement Sheet ---
wsa = wbs.create_sheet("Judge Agreement")
ahdrs = ['Model','Judge','Avg Found','Avg Total','Avg %','Prompts Judged']
for i, h in enumerate(ahdrs, 1):
    c = wsa.cell(1, i, h); c.font = hfont; c.fill = hfill; c.border = bdr; c.alignment = ctr

arow = 2
for mn in available:
    if mn not in done: continue
    for jn in JUDGE_NAMES:
        founds, totals = [], []
        for pid, d in done[mn].items():
            jd = d.get("judges", {}).get(jn, {})
            if isinstance(jd.get("found"), (int, float)) and jd["found"] != "-":
                founds.append(jd["found"])
                totals.append(jd["total"])
        
        n = len(founds)
        if n > 0:
            af = round(sum(founds)/n, 2)
            at = round(sum(totals)/n, 2)
            ap = round(af/max(at,1)*100, 1)
        else:
            af = at = ap = 0
        
        wsa.cell(arow,1,mn).border=bdr
        wsa.cell(arow,2,jn).border=bdr
        wsa.cell(arow,3,af).border=bdr; wsa.cell(arow,3).alignment=ctr
        wsa.cell(arow,4,at).border=bdr; wsa.cell(arow,4).alignment=ctr
        wsa.cell(arow,5,ap).border=bdr; wsa.cell(arow,5).alignment=ctr
        wsa.cell(arow,6,n).border=bdr; wsa.cell(arow,6).alignment=ctr
        arow += 1

for col, w in {1:12, 2:10, 3:10, 4:10, 5:8, 6:14}.items():
    wsa.column_dimensions[chr(64+col)].width = w

wss.freeze_panes = 'A2'
wsa.freeze_panes = 'A2'

for col, w in {1:12,2:10,3:10,4:10,5:10,6:10,7:12,8:8,9:8,10:8,11:8,12:8}.items():
    wss.column_dimensions[chr(64+col)].width = w

wbs.save("ChemEng_Bench_Summary_v3.xlsx")
print(f"\n  Saved: ChemEng_Bench_Summary_v3.xlsx")

# --- Stats Report ---
print(f"\n{'='*60}")
print("EVALUATION REPORT")
print(f"{'='*60}")
for mn in available:
    if mn not in done: continue
    nj_counts = Counter()
    for pid, d in done[mn].items():
        nj_counts[d.get("num_judges",0)] += 1
    print(f"\n  {mn}:")
    for k in sorted(nj_counts.keys()):
        status = "OK" if k >= MIN_JUDGES else "NEEDS REDO"
        print(f"    {k} judges: {nj_counts[k]} prompts  [{status}]")

print(f"\n{'='*60}")
print(f"ALL DONE! Weights: R1={W_ROUGE1} RL={W_ROUGEL} Cos={W_COSINE} Elem={W_ELEMENT}")
print(f"{'='*60}")
